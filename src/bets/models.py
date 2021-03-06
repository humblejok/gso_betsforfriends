from django.contrib.auth.models import User
from django.db import models
from django.template import loader
from openpyxl.reader.excel import load_workbook
from seq_common.utils import classes, dates
from datetime import datetime as dt

import datetime
import logging
import os
import traceback
from django.template.context import Context
from gso_betsforfriends.settings import RESOURCES_MAIN_PATH, STATICS_PATH, STATIC_TEMPLATES_PATH
from django.db.models.aggregates import Sum
from suds.client import Client
from django.db.models.fields import FieldDoesNotExist

LOGGER = logging.getLogger(__name__)

def setup():
    populate_attributes_from_xlsx('bets.models.Attributes', os.path.join(RESOURCES_MAIN_PATH,'Repository Setup.xlsx'))
    generate_attributes()
    populate_model_from_xlsx('bets.models.Participant', os.path.join(RESOURCES_MAIN_PATH,'Repository Setup.xlsx'))
    populate_model_from_xlsx('bets.models.Match', os.path.join(RESOURCES_MAIN_PATH,'Repository Setup.xlsx'))
    populate_model_from_xlsx('bets.models.BettableEvent', os.path.join(RESOURCES_MAIN_PATH,'Repository Setup.xlsx'))

def setup_attributes_only():
    populate_attributes_from_xlsx('bets.models.Attributes', os.path.join(RESOURCES_MAIN_PATH,'Repository Setup.xlsx'))
    generate_attributes()

def populate_model(model_name, file_name):
    populate_model_from_xlsx(model_name, os.path.join(RESOURCES_MAIN_PATH, file_name))

def update_model(model_name, file_name, keys, fields):
    update_model_from_xlsx(model_name, os.path.join(RESOURCES_MAIN_PATH, file_name), keys, fields)

def generate_attributes():
    all_types = Attributes.objects.all().order_by('type').distinct('type')
    for a_type in all_types:
        all_elements = Attributes.objects.filter(type=a_type.type, active=True)
        context = Context({"selection": all_elements})
        template = loader.get_template('rendition/attributes_option_renderer.html')
        rendition = template.render(context)
        # TODO Implement multi-langage
        outfile = os.path.join(STATICS_PATH, 'attributes', a_type.type + '_en.html')
        templatefile = os.path.join(STATIC_TEMPLATES_PATH, a_type.type + '_en.html')
        with open(outfile,'w') as o:
            with open(templatefile, 'w') as t:
                o.write(rendition.encode('utf-8'))
                t.write(rendition.encode('utf-8'))
            
def generate_events():
    start_date = dates.AddDay(datetime.date.today(),-14)
    start_date_time = dt.combine(start_date, dt.min.time())
    all_dates = Match.objects.filter(when__gte=start_date_time).order_by('when').dates('when','day')
    template = loader.get_template('rendition/events.html')
    for a_date in all_dates:
        events = BettableEvent.objects.filter(end_date__gte=start_date)
        for event in events:
            matchs = event.matchs.filter(when__gte=dt.combine(a_date, dt.min.time()), when__lte=dt.combine(a_date, dt.max.time())).order_by('when')
            context = Context({"events": matchs})
            rendition = template.render(context)
            try:
                os.makedirs(os.path.join(STATICS_PATH,'events', str(event.id)))
            except:
                None # Folder already exists
            outfile = os.path.join(STATICS_PATH,'events', str(event.id), a_date.strftime('%Y-%m-%d') + '_en.html')
            with open(outfile,'w') as o:
                o.write(rendition.encode('utf-8'))
            
def generate_matchs(match_ids=None):
    start_date = dates.AddDay(datetime.date.today(),-14)
    start_date_time = dt.combine(start_date, dt.min.time())
    all_dates = Match.objects.filter(when__gte=start_date_time).order_by('when').dates('when','day')
    template = loader.get_template('rendition/matchs.html')
    for a_date in all_dates:
        events = BettableEvent.objects.filter(end_date__gte=start_date)
        for event in events:
            matchs = event.matchs.filter(when__gte=dt.combine(a_date, dt.min.time()), when__lte=dt.combine(a_date, dt.max.time())).order_by('when')
            context = Context({"events": matchs})
            rendition = template.render(context)
            try:
                os.makedirs(os.path.join(STATICS_PATH,'matchs', str(event.id)))
            except:
                None # Folder already exists
            outfile = os.path.join(STATICS_PATH,'matchs', str(event.id), a_date.strftime('%Y-%m-%d') + '_en.html')
            with open(outfile,'w') as o:
                o.write(rendition.encode('utf-8'))

def populate_attributes_from_xlsx(model_name, xlsx_file):
    model = classes.my_class_import(model_name)
    workbook = load_workbook(xlsx_file)
    sheet = workbook.get_sheet_by_name(name=model.__name__)
    row_index = 1
    # Reading header
    header = []
    for column_index in range(1, sheet.get_highest_column()+1):
        value = sheet.cell(row = row_index, column=column_index).value
        if value!=None:
            header.append(value if value!='' else header[-1])
        else:
            break
    LOGGER.info('Using header:' + str(header))
    row_index += 1
    while row_index<=sheet.get_highest_row():
        if model.objects.filter(identifier=sheet.cell(row = row_index, column = 1).value).exists():
            instance = model.objects.get(identifier=sheet.cell(row = row_index, column = 1).value)
        else:
            instance = model()
        for i in range(0,len(header)):
            value = sheet.cell(row = row_index, column = i+1).value
            setattr(instance, header[i], value)
        if instance.identifier==None:
            break
        else:
            instance.save()
        row_index += 1

def update_model_from_xlsx(model_name, xlsx_file, keys, fields):
    LOGGER.info("Loading data in " + model_name)
    model = classes.my_class_import(model_name)
    workbook = load_workbook(xlsx_file)
    sheet = workbook.get_sheet_by_name(name=model.__name__)
    row_index = 1
    # Reading header
    header = []
    for column_index in range(1, sheet.get_highest_column() + 1):
        value = sheet.cell(row = row_index, column=column_index).value
        if value!=None:
            header.append(value if value!='' else header[-1])
        else:
            break
    LOGGER.info('Using header:' + str(header))
    row_index += 1
    while row_index<=sheet.get_highest_row():
        filtering_by = {}
        for key in keys:
            if model._meta.get_field(key).get_internal_type()=='ForeignKey':
                filtering_by[key + '__name'] = sheet.cell(row = row_index, column = header.index(key, ) + 1).value
            else:
                filtering_by[key] = sheet.cell(row = row_index, column = header.index(key, ) + 1).value
        by_identifier = model.objects.filter(**filtering_by)
        if by_identifier.exists():
            instance = by_identifier[0]
            for i in range(0,len(header)):
                if header[i] in fields:
                    value = sheet.cell(row = row_index, column = i + 1).value
                    field_info = Attributes()
                    field_info.short_name = header[i]
                    field_info.name = header[i]
                    instance.set_attribute('excel', field_info, value)
            instance.finalize()
            if instance.name==None:
                break
            else:
                instance.save()
        else:
            LOGGER.error("Could not find model with " + str(by_identifier))
        row_index += 1

def populate_model_from_xlsx(model_name, xlsx_file):
    LOGGER.info("Loading data in " + model_name)
    model = classes.my_class_import(model_name)
    workbook = load_workbook(xlsx_file)
    sheet = workbook.get_sheet_by_name(name=model.__name__)
    row_index = 1
    # Reading header
    header = []
    for column_index in range(1, sheet.get_highest_column() + 1):
        value = sheet.cell(row = row_index, column=column_index).value
        if value!=None:
            header.append(value if value!='' else header[-1])
        else:
            break
    LOGGER.info('Using header:' + str(header))
    row_index += 1
    while row_index<=sheet.get_highest_row():
        instance = model()
        for i in range(0,len(header)):
            value = sheet.cell(row = row_index, column = i + 1).value
            field_info = Attributes()
            field_info.short_name = header[i]
            field_info.name = header[i]
            instance.set_attribute('excel', field_info, value)
        instance.finalize()
        if instance.name==None:
            break
        else:
            instance.save()
        row_index += 1
        
def rank_list(rank_list):
    ordered = sorted(rank_list, key=lambda x: x.overall_score, reverse=True)
    ranking = 1
    for rank in ordered:
        rank.rank = ranking
        ranking += 1
    [rank.save() for rank in ordered]

def compute_event_ranking():
    users = User.objects.filter(is_active=True)
    today = dates.AddDay(datetime.date.today(),-1)
    events = BettableEvent.objects.filter(end_date__gte=today)
    events_ranking = {}
    for event in events:
        LOGGER.info("Working on event " + unicode(event.name))
        EventRanking.objects.filter(event__id=event.id).delete()
        for user in users:
            LOGGER.info("\tWorking on user " + unicode(user.username))
            for match in event.matchs.filter(result__isnull=False):
                score = 0
                LOGGER.info("\t\tWorking on match " + unicode(match.name))
                bet = Bet.objects.filter(match__id=match.id, owner__id=user.id)
                winner = match.get_winner()
                LOGGER.info("\t\tWinner is " + unicode(winner))
                if bet.exists():
                    bet = bet[0]
                    score = bet.get_score()
                    if not events_ranking.has_key(event.id):
                        events_ranking[event.id] = {}
                    if not events_ranking[event.id].has_key(user.id):
                        rank = EventRanking()
                        rank.event = event
                        rank.owner = user
                        rank.overall_score = 0
                        rank.rank = None
                        events_ranking[event.id][user.id] = rank
                        LOGGER.info("\t\tResult: FINAL CREATION = " + str(events_ranking[event.id][user.id].overall_score))
                    events_ranking[event.id][user.id].overall_score += score
                    LOGGER.info("\t\tResult: FINAL score = " + str(events_ranking[event.id][user.id].overall_score))
    for event_id in events_ranking.keys():
        rank_list(events_ranking[event_id].values())

def compute_overall_ranking():
    users = User.objects.filter(is_active=True)
    ranks = []
    for user in users:
        LOGGER.info("\tWorking on user " + str(user.username))
        global_score = EventRanking.objects.filter(owner__id=user.id).aggregate(Sum('overall_score'))['overall_score__sum']
        ranking = UserRanking.objects.filter(owner__id=user.id, group=None)
        if ranking.exists():
            ranking = ranking[0]
        else:
            ranking = UserRanking()
            ranking.owner = user
            ranking.group = None
            ranking.overall_score = 0
            ranking.rank = None
            ranking.save()
        ranking.overall_score = global_score if global_score!=None else 0
        ranking.save()
        ranks.append(ranking)
        rank_list(ranks)

class CoreModel(models.Model):

    many_fields = {}
    
    def finalize(self):
        self.save()
        # TODO Loop on many to many only
        for field_name in self._meta.get_all_field_names():
            try:
                if self._meta.get_field(field_name).get_internal_type()=='ManyToManyField':
                    if self.many_fields.has_key(field_name):
                        values = list(self.many_fields[field_name])
                        setattr(self, field_name, values)
            except FieldDoesNotExist:
                None
        self.save()

    def get_editable_fields(self):
        values = []
        for field in self.get_fields():
            if self._meta.get_field(field).get_internal_type()!='ManyToManyField':
                values.append(field)
        return values
        
    def get_associable_field(self):
        values = []
        for field in self.get_fields():
            if self._meta.get_field(field).get_internal_type()=='ManyToManyField':
                values.append(field)
        return values        

    def get_fields(self):
        return []
    
    def get_identifier(self):
        for field in self.get_fields():
            if field=='name':
                return 'name'
        return 'id'
    
    def list_values(self):
        values = []
        for field in self.get_fields():
            LOGGER.debug(self.__class__.__name__ + ' * ' + field)
            if field in self._meta.get_all_field_names():
                if self._meta.get_field(field).get_internal_type()=='ManyToManyField' and getattr(self,field)!=None:
                    values.append(str([e.list_values() for e in list(getattr(self,field).all())]))
                elif self._meta.get_field(field).get_internal_type()=='ForeignKey' and getattr(self,field)!=None:
                    values.append(getattr(self,field).get_value())
                else:
                    values.append(str(getattr(self,field)))
            else:
                # Generic foreign key
                values.append(getattr(self,field).get_value())
        return values
    
    def get_value(self):
        if self.get_identifier()!=None:
            return getattr(self, self.get_identifier())
        else:
            return None
        
    def __unicode__(self):
        return unicode(self.get_value())
    
    def set_attribute(self, source, field_info, string_value):
        try:
            if string_value!='' and string_value!=None:
                if self._meta.get_field(field_info.short_name).get_internal_type()=='ManyToManyField':
                    if not self.many_fields.has_key(field_info.short_name):
                        self.many_fields[field_info.short_name] = []
                    foreign = self._meta.get_field(field_info.short_name).rel.to
                    elements = string_value.split(',')
                    for element in elements:
                        foreign_entity = foreign.retrieve_or_create(source, field_info.name, element)
                        print foreign_entity
                        self.many_fields[field_info.short_name].append(foreign_entity)
                elif self._meta.get_field(field_info.short_name).get_internal_type()=='DateField' or self._meta.get_field(field_info.short_name).get_internal_type()=='DateTimeField':
                    try:
                        dt = datetime.datetime.strptime(string_value,'%m/%d/%Y')
                        if self._meta.get_field(field_info.short_name).get_internal_type()=='DateField':
                            dt = datetime.date(dt.year, dt.month, dt.day)
                    except:
                        dt = string_value # This is not a String???
                    setattr(self, field_info.short_name, dt)
                elif self._meta.get_field(field_info.short_name).get_internal_type()=='ForeignKey':
                    linked_to = self._meta.get_field(field_info.short_name).rel.limit_choices_to
                    foreign = self._meta.get_field(field_info.short_name).rel.to
                    filtering_by_identifier = dict(linked_to)
                    filtering_by_identifier['identifier'] = string_value
                    try:
                        by_identifier = foreign.objects.filter(**filtering_by_identifier)
                    except:
                        by_identifier = None
                    filtering_by_name = dict(linked_to)
                    filtering_by_name['name'] = string_value
                    try:
                        by_name = foreign.objects.filter(**filtering_by_name)
                    except:
                        by_name = None
                    filtering_by_short = dict(linked_to)
                    filtering_by_short['short_name'] = string_value
                    try:
                        by_short = foreign.objects.filter(**filtering_by_short)
                    except:
                        by_short = None
                    if by_identifier!=None and by_identifier.exists():
                        setattr(self, field_info.short_name, by_identifier[0])
                    elif by_name!=None and by_name.exists():
                        setattr(self, field_info.short_name, by_name[0])
                    elif by_short!=None and by_short.exists():
                        setattr(self, field_info.short_name, by_short[0])
                    else:
                        dict_entry = Dictionary.objects.filter(name=linked_to['type'], auto_create=True)
                        if dict_entry.exists():
                            LOGGER.info('Creating new attribute for ' + linked_to['type'] + ' with value ' + string_value)
                            new_attribute = Attributes()
                            new_attribute.active = True
                            new_attribute.identifier = dict_entry[0].identifier + str(string_value.upper()).replace(' ', '_')
                            new_attribute.name = string_value
                            new_attribute.short_name = string_value[0:32]
                            new_attribute.type = linked_to['type']
                            new_attribute.save()
                            setattr(self, field_info.short_name, new_attribute)
                        else:
                            LOGGER.warn('Cannot find foreign key instance on ' + str(self) + '.' + field_info.short_name + ' for value [' + string_value + '] and relation ' + str(linked_to))
                else:
                    setattr(self, field_info.short_name, string_value)
        except FieldDoesNotExist:
            traceback.print_exc()
            LOGGER.error("Wrong security type for " + self.name + ", please check your settings...")
    
    class Meta:
        ordering = ['id']

class Group(CoreModel):
    name = models.CharField(max_length=128)
    owners = models.ManyToManyField(User, related_name='group_owners_rel')
    members = models.ManyToManyField(User, related_name='group_members_rel')
    event = models.ForeignKey('BettableEvent', related_name='group_event')
    
    def get_fields(self):
        return ['name','owners','members', 'event']
    
    class Meta:
        ordering = ['name']

class EventRanking(CoreModel):
    owner = models.ForeignKey(User, related_name='ranking_event_user')
    event = models.ForeignKey('BettableEvent', related_name='ranking_event')
    overall_score = models.IntegerField(default=0)
    rank = models.IntegerField(null=True)
    
    def get_fields(self):
        return ['owner','event','overall_score', 'rank']
    
    class Meta:
        ordering = ['rank']
    
class UserRanking(CoreModel):
    owner = models.ForeignKey(User, related_name='ranking_owner_rel')
    group = models.ForeignKey(Group, related_name='ranking_group_rel', null=True)
    overall_score = models.IntegerField(default=0)
    rank = models.IntegerField(null=True)

    def get_fields(self):
        return ['owner','group','overall_score', 'rank']
    
    class Meta:
        ordering = ['rank']

class Attributes(CoreModel):
    identifier = models.CharField(max_length=128)
    name = models.CharField(max_length=128)
    short_name = models.CharField(max_length=32)
    type = models.CharField(max_length=64)
    active = models.BooleanField()
    
    def get_fields(self):
        return ['identifier','name','short_name','type','active']
    
    class Meta:
        ordering = ['name']
        
class Dictionary(CoreModel):
    identifier = models.CharField(max_length=128)
    name = models.CharField(max_length=128)
    auto_create = models.BooleanField()
    
    def get_fields(self):
        return ['identifier','name','auto_create']
    
    class Meta:
        ordering = ['name']
        
class Participant(CoreModel):
    name = models.CharField(max_length=256)
    country = models.ForeignKey(Attributes, limit_choices_to={'type':'country_iso2'}, related_name='participant_country_rel', null=True)
    sport = models.ForeignKey(Attributes, limit_choices_to={'type':'sport'}, related_name='participant_sport_rel', null=True)

    def get_fields(self):
        return super(Participant, self).get_fields() + ['name','country','sport']

    @staticmethod
    def retrieve_or_create(source, key, value):
        translation = Attributes.objects.filter(active=True, name=key, type=source.lower() + '_translation')
        if translation.exists():
            translation = translation[0].short_name
        else:
            translation = key
        participant = Participant.objects.filter(name=value)
        if participant.exists():
            return participant[0]
        else:
            return None

class Score(CoreModel):
    name = models.CharField(max_length=256)
    first = models.IntegerField(default=0)
    second = models.IntegerField(default=0)
    
    def get_fields(self):
        return super(Score, self).get_fields() + ['name', 'first','second']
            
class Match(CoreModel):
    name = models.CharField(max_length=256)
    type = models.ForeignKey(Attributes, limit_choices_to={'type':'match_type'}, related_name='match_type_rel')
    when = models.DateTimeField()
    first = models.ForeignKey(Participant, related_name='match_first_part_rel')
    second = models.ForeignKey(Participant, related_name='match_second_part_rel')
    result = models.ForeignKey(Score, related_name='match_score_rel', null=True)
    
    def display_score(self):
        if self.result:
            display_score = str(self.result.first) + " - " + str(self.result.second)
            return display_score
        return "Aucun" 
    
    def get_winner(self):
        winner = None
        
        if self.result!=None and self.result.first!=self.result.second:
            winner = self.first if self.result.first>self.result.second else self.second
        return winner
    
    def get_fields(self):
        return super(Match, self).get_fields() + ['name','type','when','first','second','result']
    
    @staticmethod
    def retrieve_or_create(source, key, value):
        translation = Attributes.objects.filter(active=True, name=key, type=source.lower() + '_translation')
        if translation.exists():
            translation = translation[0].short_name
        else:
            translation = key
        match = Match.objects.filter(name=value)
        if match.exists():
            return match[0]
        else:
            return None

class Bet(CoreModel):
    owner = models.ForeignKey(User, related_name='bet_owner_rel')
    when = models.DateTimeField()
    match = models.ForeignKey(Match, related_name='bet_match_rel')
    winner = models.ForeignKey(Participant, related_name='bet_winner_rel', null=True)
    result = models.ForeignKey(Score, related_name='bet_score_rel')
    amount = models.IntegerField(null=True)

    def get_score(self):
        score = 0
        winner = self.match.get_winner()
        if self.match.result!=None and self.result!=None:
            if self.winner==None and winner==None:
                score += 1
                LOGGER.info("\t\tResult: Even found = " + str(score))
            elif self.winner!=None and winner!=None:
                score += 1 if self.winner.id==winner.id else 0
                LOGGER.info("\t\tResult: Winner/Looser = " + str(score))
            score += 3 if self.result.first==self.match.result.first and self.result.second==self.match.result.second else 0
        LOGGER.info("\t\tResult: Score deduction = " + str(score))
        return score

    def get_fields(self):
        return super(Bet, self).get_fields() + ['owner','when','match','winner','result']

class PointsSetup(CoreModel):
    category = models.ForeignKey(Attributes, limit_choices_to={'type':'match_type'}, related_name='points_category_rel')
    points = models.IntegerField(default=3)
    use_quotes = models.BooleanField(default=False)
    
class WinnerSetup(CoreModel):
    group = models.ForeignKey(Group, related_name='winner_setup_group_rel', null=True)
    setup = models.ManyToManyField(PointsSetup, related_name='winner_setup_points_rel')
    
class Winner(CoreModel):
    owner = models.ForeignKey(User, related_name='winner_owner_rel')
    event = models.ForeignKey("BettableEvent", related_name='winner_event_rel')
    category = models.ForeignKey(Attributes, limit_choices_to={'type':'match_type'}, related_name='winner_category_rel')
    participants = models.ManyToManyField(Participant, related_name='winner_participant_rel')
    
    def get_score(self):
        return 0
    
class BettableEvent(CoreModel):
    name = models.CharField(max_length=256)
    sport = models.ForeignKey(Attributes, limit_choices_to={'type':'sport'}, related_name='event_sport_rel', null=True)
    start_date = models.DateField()
    end_date = models.DateField()
    participants = models.ManyToManyField(Participant, related_name='event_participant_rel')
    matchs = models.ManyToManyField(Match, related_name='event_match_rel')
    
    def get_fields(self):
        return super(BettableEvent, self).get_fields() + ['name','sport','start_date','end_date','participants','matchs']

class Provider(CoreModel):
    name = models.CharField(max_length=256)
    event = models.ForeignKey('BettableEvent', related_name='provider_event_rel')
    source = models.ForeignKey(Attributes, limit_choices_to={'type':'provider_source'}, related_name='provider_source_rel')
    ws_url = models.CharField(max_length=1024, null=True)
    ws_all_method = models.CharField(max_length=1024, null=True)
    ws_all_arguments = models.CharField(max_length=1024, null=True)
    ws_unique_method = models.CharField(max_length=1024, null=True)
    ws_unique_arguments = models.CharField(max_length=1024, null=True)
    
    def get_fields(self):
        return super(Provider, self).get_fields() + ['name','event', 'source','ws_url','ws_all_method','ws_all_arguments','ws_unique_method','ws_unique_arguments']
    
    def initialize_tables(self):
        for match in self.event.matchs.all():
            LOGGER.info("Working on " + match.name)
            mapping = ProviderMapping.objects.filter(local__id=match.id, provider__id=self.id)
            if mapping.exists():
                LOGGER.info("\tAlready mapped, skipping!")
            else:
                mapping = ProviderMapping()
                mapping.local = match
                mapping.name = match.name
                mapping.local_model = 'Match'
                mapping.provider = self
                mapping.save()
        for participant in self.event.participants.all():
            LOGGER.info("Working on " + participant.name)
            mapping = ProviderMapping.objects.filter(local__id=participant.id, provider__id=self.id)
            if mapping.exists():
                LOGGER.info("\tAlready mapped, skipping!")
            else:
                mapping = ProviderMapping()
                mapping.local = participant
                mapping.local_model = 'Participant'
                mapping.name = participant.name
                mapping.provider = self
                mapping.save()
                
    def get_all_matches_info(self):
        if self.name=='footballpool.dataaccess.eu':
            client = Client(self.ws_url)
            all_info = getattr(client.service,self.ws_all_method)()
            for info in all_info.tGameInfo:
                valid_score = info.sResult!='U'
                if valid_score:
                    match_id = info.iId
                    first_id = info.Team1.iId
                    second_id = info.Team2.iId
                    score = info.sScore
                    match = Match.objects.get(id=ProviderMapping.objects.get(target_id=match_id, local_model='Match', provider__id=self.id).local.id)
                    first = Participant.objects.get(id=ProviderMapping.objects.get(target_id=first_id, local_model='Participant', provider__id=self.id).local.id)
                    second = Participant.objects.get(id=ProviderMapping.objects.get(target_id=second_id, local_model='Participant', provider__id=self.id).local.id)
                    score_info = score.split('-')
                    if match.first != first:
                        first = int(score_info[1])
                        second = int(score_info[0])
                    else:
                        first = int(score_info[0])
                        second = int(score_info[1])
                    LOGGER.info(match.name)
                    LOGGER.info(score + " => " + str(first) + "-" + str(second))
                    if match.result!=None:
                        result = Score.objects.get(id=match.result.id)
                    else:
                        result = Score()
                        result.name = match.name
                        result.save()
                        match.result = result
                        match.save()
                    result.first = first
                    result.second = second
                    result.save()
                    
    
class ProviderMapping(CoreModel):
    local = models.ForeignKey(CoreModel, related_name='provider_local_rel')
    name = models.CharField(max_length=1024, null=True)
    local_model = models.CharField(max_length=1024, null=True)
    provider = models.ForeignKey(Provider, related_name='provider_rel')
    target_name = models.CharField(max_length=1024, null=True)
    target_id = models.IntegerField(null=True)
    
    def get_fields(self):
        return super(ProviderMapping, self).get_fields() + ['local', 'name', 'local_model', 'provider', 'target_name', 'target_id']